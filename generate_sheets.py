import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# Set target file directory and name
target_dir = r"C:\Users\Administrator\Documents\KKN\WEB KKN"
os.makedirs(target_dir, exist_ok=True)
excel_path = os.path.join(target_dir, "Absensi KKN Ambarawa Timur.xlsx")
csv_path = os.path.join(target_dir, "Absensi KKN Ambarawa Timur.csv")

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Absensi KKN"

# Enable grid lines visibility explicitly
ws.views.sheetView[0].showGridLines = True

# Title Row
ws.merge_cells("A1:G1")
title_cell = ws["A1"]
title_cell.value = "REKAPITULASI ABSENSI KKN AMBARAWA TIMUR"
title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
title_cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Royal Blue
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

# Subtitle / Info Row
ws.merge_cells("A2:G2")
sub_cell = ws["A2"]
sub_cell.value = "Lokasi: Kecamatan Ambarawa Timur | Periode: 2026"
sub_cell.font = Font(name="Calibri", size=10, italic=True, color="4B5563")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# Table Headers (Row 4)
headers = ["Timestamp", "Nama", "Tanggal", "Waktu", "Status", "Keterangan", "Foto"]
header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Modern Blue
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB")
)

header_row = 4
ws.row_dimensions[header_row].height = 28

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Single Sample Data (Sisakan 1 data dummy sebagai contoh)
sample_data = [
    ["2026-07-24 07:30:15", "Ahmad Fauzi", "2026-07-24", "07:30", "Hadir", "Piket Posko Pagi", "foto_ahmad_20260724.jpg"]
]

# Fill Sample Data
start_row = 5
zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for r_idx, row_data in enumerate(sample_data, start=start_row):
    ws.row_dimensions[r_idx].height = 22
    row_fill = zebra_fill if r_idx % 2 == 0 else white_fill
    
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.value = val
        cell.fill = row_fill
        cell.font = Font(name="Calibri", size=11)
        cell.border = thin_border
        
        if c_idx in [1, 3, 4, 5]:  # Timestamp, Tanggal, Waktu, Status
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

# Prepare empty formatted rows up to row 50 for future entries
for r_idx in range(start_row + len(sample_data), 51):
    ws.row_dimensions[r_idx].height = 22
    row_fill = zebra_fill if r_idx % 2 == 0 else white_fill
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.fill = row_fill
        cell.font = Font(name="Calibri", size=11)
        cell.border = thin_border
        if c_idx in [1, 3, 4, 5]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

# Data Validation for Status Column (Column E)
dv = DataValidation(type="list", formula1='"Hadir,Izin,Sakit,Alpa"', allow_blank=True)
dv.error = "Pilih status dari daftar (Hadir, Izin, Sakit, Alpa)"
dv.errorTitle = "Status Tidak Valid"
dv.prompt = "Pilih status absensi"
dv.promptTitle = "Status"
ws.add_data_validation(dv)
dv.add(f"E5:E50")

# Column Width Optimization
min_widths = {
    1: 22, # Timestamp
    2: 20, # Nama
    3: 15, # Tanggal
    4: 12, # Waktu
    5: 14, # Status
    6: 30, # Keterangan
    7: 25  # Foto
}

for col_idx, min_w in min_widths.items():
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = min_w

# Save Excel File
wb.save(excel_path)
print(f"Excel created successfully at: {excel_path}")

# Export CSV File (standard UTF-8 with header)
import csv
with open(csv_path, mode="w", newline="", encoding="utf-8-sig") as csv_file:
    writer = csv.writer(csv_file)
    writer.writerow(headers)
    for row in sample_data:
        writer.writerow(row)

print(f"CSV created successfully at: {csv_path}")
